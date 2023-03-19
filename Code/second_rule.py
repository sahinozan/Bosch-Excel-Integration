from customtkinter import CTkLabel, CTkOptionMenu, CTkButton, CTk, set_appearance_mode, set_default_color_theme
import pandas as pd
import openpyxl
import sys


class SecondRuleWindow(CTk):
    set_appearance_mode("Dark")
    set_default_color_theme("dark-blue")

    def __init__(self) -> None:
        super().__init__()

        self.dates = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday*"]
        self.score_labels = []
        self.score_entries = []
        self.next_week_excel_path = self.find_output_path()
        self.next_week_workbook = openpyxl.load_workbook(self.next_week_excel_path)
        self.next_week_worksheet = self.next_week_workbook['Pivot']
        self.next_week_df = pd.DataFrame()
        self.window_configuration()
        self.create_grid()

    def find_output_path(self) -> str:
        with open("output_path.txt", "r") as f:
            return f.read()

    def window_configuration(self) -> None:
        """
        Sets the title of the window, makes the window non-resizable, and makes the window close when the user clicks.
        Also sets the dimensions of the window and places it in the center of the screen.
        """
        self.title("Shift Selection")

        # make the window non-resizable (for macOS and Linux)
        if sys.platform == "darwin":
            self.resizable(False, False)

        # set the dimensions of the window
        w = 565  # width for the Tk root
        h = 395  # height for the Tk root

        # get screen width and height
        ws = self.winfo_screenwidth()  # width of the screen
        hs = self.winfo_screenheight()  # height of the screen

        # calculate x and y coordinates for the Tk root window
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)

        # set the dimensions of the screen where the window will be displayed
        self.geometry('%dx%d+%d+%d' % (w, h, x, y))

    def save_data(self) -> None:
        next_week_data = {'Gün': [date for date in self.dates]}

        for shift in range(3):
            next_week_data[f'Shift {shift + 1}'] = [
                self.score_entries[self.dates.index(date) * 3 + shift].get() for date in self.dates]

        self.next_week_df = pd.DataFrame(next_week_data)

        self.next_func()
        self.destroy()

    def next_func(self) -> None:
        last_row = self.next_week_worksheet.max_row
        matching_rows = [i for i in range(3, last_row+1)]

        shifts = []
        for row in matching_rows:
            row_shifts = []
            for col in range(4, 28):
                row_shifts.append(self.next_week_worksheet.cell(row=row, column=col).value)
            shifts.append(row_shifts)

        for i in range(len(shifts)):
            for j in range(len(shifts[i])):
                if shifts[i][j] is None:
                    shifts[i][j] = 0

        shift_list = []
        for i in range(len(self.next_week_df)):
            for j in range(len(self.next_week_df.columns)):
                if self.next_week_df.iloc[i, j] == "Var":
                    shift_list.append((i * 3) + j - 1)

        for k in range(len(shifts)):
            for i in shift_list:
                shifts[k][i - 1] = int(shifts[k][i]) + int(shifts[k][i - 1])
                shifts[k][i] = 0

        for row, lists in zip(matching_rows, shifts):
            for col, element in zip(range(13, 34), lists):
                self.next_week_worksheet.cell(row=row, column=col).value = element

        for i in range(len(shifts)):
            for j in range(len(shifts[i])):
                if shifts[i][j] == 0:
                    shifts[i][j] = None

        for row, lists in zip(matching_rows, shifts):
            for col, element in zip(range(4, 28), lists):
                self.next_week_worksheet.cell(row=row, column=col).value = element

        self.next_week_workbook.save(self.next_week_excel_path)

    def create_grid(self) -> None:
        header_labels = [CTkLabel(self, text=f"Shift {i + 1}") for i in range(3)]
        for i, header in enumerate(header_labels):
            header.grid(row=0, column=i + 1, sticky="nsew", padx=5, pady=5)

        for date in self.dates:
            date_label = CTkLabel(self, text=date)
            date_label.grid(row=self.dates.index(date) + 1, column=0, sticky="nsew", padx=15, pady=5)

            for shift in range(3):
                score_entry = CTkOptionMenu(self, values=["Yok", "Var"], anchor="center")
                score_entry.grid(row=self.dates.index(date) + 1, column=shift + 1, padx=5, pady=5, sticky="nsew")

                self.score_entries.append(score_entry)
                self.score_labels.append(date)

        save_button = CTkButton(self, text="Save", command=self.save_data)
        save_button.grid(row=len(self.dates) + 2, column=3, padx=5, pady=10)


if __name__ == "__main__":
    app = SecondRuleWindow()
    app.mainloop()
