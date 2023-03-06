# Author: Ozan Şahin

import sys
import openpyxl
import pandas as pd
from customtkinter import CTkToplevel, CTkTabview, CTkLabel, CTkOptionMenu, CTkButton
from tkinter import Tk, messagebox


class ShiftWindow(CTkToplevel):

    def __init__(self, *args, next_week_excel_path, current_week_excel_path, **kwargs):
        super().__init__(*args, **kwargs)

        self.dates = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.score_labels = []
        self.score_entries = []
        self.next_week_excel_path = next_week_excel_path
        self.current_week_excel_path = current_week_excel_path
        self.window_configuration()
        self.create_grid()
        self.next_week_workbook = openpyxl.load_workbook(self.next_week_excel_path)
        self.current_week_workbook = openpyxl.load_workbook(self.current_week_excel_path)
        self.next_week_worksheet = self.next_week_workbook['GZT-GWT']
        self.current_week_worksheet = self.current_week_workbook['GZT-GWT']
        self.next_week_df = pd.DataFrame()
        self.current_week_df = pd.DataFrame()

    def window_configuration(self) -> None:
        """
        Sets the title of the window, makes the window non-resizable, and makes the window close when the user clicks.
        Also sets the dimensions of the window and places it in the center of the screen.
        """
        self.title("Shift Selection")

        # make the window non-resizable (for macOS and Linux)
        if sys.platform == "darwin":
            self.resizable(False, False)

        # set the dimensions of the window
        w = 565  # width for the Tk root
        h = 420  # height for the Tk root

        # get screen width and height
        ws = self.winfo_screenwidth()  # width of the screen
        hs = self.winfo_screenheight()  # height of the screen

        # calculate x and y coordinates for the Tk root window
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)

        # set the dimensions of the screen where the window will be displayed
        self.geometry('%dx%d+%d+%d' % (w, h, x, y))

    def show_info(self, message: str, system_exit=True) -> None:
        """
        Shows an information message.

        Args:
            message: The error message that will be displayed. (e.g. "You have not selected the first Excel file!")
            system_exit: Whether the program should be closed after the error message has been displayed.
        """
        warning_window = Tk()

        warning_window.withdraw()
        messagebox.showinfo(title="Information", message=message, icon="info", parent=self)
        warning_window.destroy()
        if system_exit:
            sys.exit(0)

    def save_data(self):
        next_week_data = {'Gün': [date for date in self.dates]}
        current_week_data = {'Gün': [date for date in self.dates]}
        next_week_score_entries = self.score_entries[21:]
        current_week_score_entries = self.score_entries[:21]

        for shift in range(3):
            next_week_data[f'Shift {shift + 1}'] = [next_week_score_entries[self.dates.index(date) * 3 + shift].get()
                                                    for date in self.dates]
            current_week_data[f'Shift {shift + 1}'] = [
                current_week_score_entries[self.dates.index(date) * 3 + shift].get() for date in self.dates]

        self.next_week_df = pd.DataFrame(next_week_data)
        self.current_week_df = pd.DataFrame(current_week_data)

        self.next_func()
        self.current_func()

        if next_week_score_entries[0].get() == "Var":
            self.show_info(message="""Gelecek hafta planında Pazartesi 1. Vardiya için işlem gerçekleştirilemiyor. 
            Lütfen bu vardiyada üretilmesi planlanan boruları ve adetleri, bu hafta pazar günü 3. vardiyaya 
            (veya en son hangi vardiyada üretim olacaksa) aktarınız.""", system_exit=False)

        self.destroy()

    def next_func(self):

        last_row = self.next_week_worksheet.max_row
        matching_rows = [i for i in range(1, last_row + 1) if
                         self.next_week_worksheet.cell(row=i, column=8).value and str(
                             self.next_week_worksheet.cell(row=i, column=8).value).startswith(
                             "7")]

        shifts = []
        for row in matching_rows:
            row_shifts = []
            for col in range(13, 34):
                row_shifts.append(self.next_week_worksheet.cell(row=row, column=col).value)
            shifts.append(row_shifts)

        for i in range(len(shifts)):
            for j in range(len(shifts[i])):
                if shifts[i][j] is None:
                    shifts[i][j] = 0

        shift_list = []
        for i in range(len(self.next_week_df)):
            for j in range(len(self.next_week_df.columns)):
                if self.next_week_df.iloc[i, j] == "Var":
                    shift_list.append((i * 3) + j - 1)

        for k in range(len(shifts)):
            for i in shift_list:
                shifts[k][i - 1] = int(shifts[k][i]) + int(shifts[k][i - 1])
                shifts[k][i] = 0

        for row, lists in zip(matching_rows, shifts):
            for col, element in zip(range(13, 34), lists):
                self.next_week_worksheet.cell(row=row, column=col).value = element

        self.next_week_workbook.save(self.next_week_excel_path)

    def current_func(self):
        last_row = self.current_week_worksheet.max_row
        matching_rows = [i for i in range(1, last_row + 1) if
                         self.current_week_worksheet.cell(row=i, column=8).value and str(
                             self.current_week_worksheet.cell(row=i, column=8).value).startswith(
                             "7")]

        shifts = []
        for row in matching_rows:
            row_shifts = []
            for col in range(13, 34):
                row_shifts.append(self.current_week_worksheet.cell(row=row, column=col).value)
            shifts.append(row_shifts)

        for i in range(len(shifts)):
            for j in range(len(shifts[i])):
                if shifts[i][j] is None:
                    shifts[i][j] = 0

        shift_list = []
        for i in range(len(self.current_week_df)):
            for j in range(len(self.current_week_df.columns)):
                if self.current_week_df.iloc[i, j] == "Var":
                    shift_list.append((i * 3) + j - 1)

        for k in range(len(shifts)):
            for i in shift_list:
                shifts[k][i - 1] = int(shifts[k][i]) + int(shifts[k][i - 1])
                shifts[k][i] = 0

        for row, lists in zip(matching_rows, shifts):
            for col, element in zip(range(13, 34), lists):
                self.current_week_worksheet.cell(row=row, column=col).value = element

        self.current_week_workbook.save(self.current_week_excel_path)

    def create_grid(self):
        current_week_tab_name = "Current Week"
        next_week_tab_name = "Next Week"

        current_week_identifier = self.current_week_excel_path.split("/")[-1].split(".")[0]
        next_week_identifier = self.next_week_excel_path.split("/")[-1].split(".")[0]

        tabview = CTkTabview(self)
        tabview.pack(fill="both", expand=True, pady=(0, 18))

        tabview.add(current_week_tab_name)
        tabview.add(next_week_tab_name)

        tab_names = [current_week_tab_name, next_week_tab_name]
        week_identifiers = [current_week_identifier, next_week_identifier]

        for index, tab_name in enumerate(tab_names):
            week_identifier = CTkLabel(tabview.tab(tab_name), text=week_identifiers[index])
            week_identifier.grid(row=0, column=0)

            header_labels = [CTkLabel(tabview.tab(tab_name), text=f"Shift {i + 1}") for i in range(3)]
            for i, header in enumerate(header_labels):
                header.grid(row=0, column=i + 1, sticky="nsew", padx=5, pady=5)

            for date in self.dates:
                date_label = CTkLabel(tabview.tab(tab_name), text=date)
                date_label.grid(row=self.dates.index(date) + 1, column=0, sticky="nsew", padx=15, pady=5)

                for shift in range(3):
                    score_entry = CTkOptionMenu(tabview.tab(tab_name), values=["Yok", "Var"], anchor="center")
                    score_entry.grid(row=self.dates.index(date) + 1, column=shift + 1, padx=5, pady=5, sticky="nsew")

                    self.score_entries.append(score_entry)
                    self.score_labels.append(date)

            save_button = CTkButton(tabview.tab(tab_name), text="Save", command=self.save_data)
            save_button.grid(row=len(self.dates) + 2, column=3, padx=5, pady=10)
