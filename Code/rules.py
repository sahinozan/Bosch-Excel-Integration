# Author: Ozan Şahin

import openpyxl


def third_rule(input_excel_path):
    workbook = openpyxl.load_workbook(input_excel_path)
    worksheet = workbook['Pivot']

    last_row = worksheet.max_row
    all_matching_rows = [i for i in range(1, last_row + 1) if worksheet.cell(row=i, column=3).value and
                         (str(worksheet.cell(row=i, column=3).value).startswith("Hidrolik Borusu 2") or
                          "Hidrolik Borusu *" in str(worksheet.cell(row=i, column=3).value) or
                          "Yedek Parça *" in str(worksheet.cell(row=i, column=3).value))]

    all_shifts = []

    for row in all_matching_rows:
        all_row_shifts = []
        for col in range(4, 28):
            all_row_shifts.append(worksheet.cell(row=row, column=col).value)
        all_shifts.append(all_row_shifts)

    for k in range(len(all_shifts)):
        for i in range(1, 24):
            all_shifts[k][i - 1] = all_shifts[k][i]
            all_shifts[k][i] = None

    for row, lists in zip(all_matching_rows, all_shifts):
        for col, element in zip(range(4, 28), lists):
            worksheet.cell(row=row, column=col).value = element

    matching_rows = [i for i in range(1, last_row + 1) if worksheet.cell(row=i, column=1).value and
                     ("Hat 3" in str(worksheet.cell(row=i, column=1).value) or
                      "Hat 4" in str(worksheet.cell(row=i, column=1).value)) and
                     "Hidrolik Borusu" in str(worksheet.cell(row=i, column=3).value)]

    shifts = []

    for row in matching_rows:
        row_shifts = []
        for col2 in range(4, 28):
            row_shifts.append(worksheet.cell(row=row, column=col2).value)
        shifts.append(row_shifts)

    for k2 in range(len(shifts)):
        for i2 in range(1, 24):
            shifts[k2][i2 - 1] = shifts[k2][i2]
            shifts[k2][i2] = None

    for row, lists in zip(matching_rows, shifts):
        for col, element in zip(range(4, 28), lists):
            worksheet.cell(row=row, column=col).value = element

    workbook.save(input_excel_path)

    general_sheet = workbook["Genel"]

    for row in general_sheet.iter_rows():
        for cell in row:
            if cell.value == 0:
                cell.value = None

    workbook.save(input_excel_path)

    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        ws.delete_cols(27, 3)

    workbook.save(input_excel_path)


def first_rule(input_excel_path):
    workbook = openpyxl.load_workbook(input_excel_path)
    worksheet = workbook['GZT-GWT']

    last_row = worksheet.max_row

    all_matching_rows = [i for i in range(1, last_row + 1) if worksheet.cell(row=i, column=1).value and
                         ("Hat 1" in str(worksheet.cell(row=i, column=1).value) or
                          "Hat 2" in str(worksheet.cell(row=i, column=1).value) or
                          "Hat 4" in str(worksheet.cell(row=i, column=1).value) or
                          "Hat 5" in str(worksheet.cell(row=i, column=1).value) or
                          "Hat 6" in str(worksheet.cell(row=i, column=1).value)) and
                         str(worksheet.cell(row=i, column=8).value).startswith("7")]

    all_shifts = []

    for row in all_matching_rows:
        row_vardiya = []
        for col in range(13, 34):
            row_vardiya.append(worksheet.cell(row=row, column=col).value)
        all_shifts.append(row_vardiya)

    for i in range(len(all_shifts)):
        for j in range(len(all_shifts[i])):
            if all_shifts[i][j] is None:
                all_shifts[i][j] = 0

    for k in range(len(all_shifts)):
        for i in range(0, 21):
            if all_shifts[k][i] < 70:
                all_shifts[k][i - 1] = int(all_shifts[k][i - 1]) + int(all_shifts[k][i])
                all_shifts[k][i] = 0
            else:
                all_shifts[k][i] = int(all_shifts[k][i]) - 40
                all_shifts[k][i - 1] = int(all_shifts[k][i - 1]) + 40

    for k in range(len(all_shifts)):
        for i in range(0, 21):
            if 30 > all_shifts[k][i] > 0 and all_shifts[k][i + 1] > 0:
                all_shifts[k][i + 1] = int(all_shifts[k][i + 1]) + int(all_shifts[k][i])
                all_shifts[k][i] = 0

    for row, lists in zip(all_matching_rows, all_shifts):
        for col, element in zip(range(13, 34), lists):
            worksheet.cell(row=row, column=col).value = element

    matching_rows = [i for i in range(1, last_row + 1) if worksheet.cell(row=i, column=1).value and
                     ("Hat 3" in str(worksheet.cell(row=i, column=1).value)) and
                     str(worksheet.cell(row=i, column=8).value).startswith("7")]

    shifts = []

    for row2 in shifts:
        row_vardiya2 = []
        for col2 in range(13, 34):
            row_vardiya2.append(worksheet.cell(row=row2, column=col2).value)
        shifts.append(row_vardiya2)

    for i in range(len(shifts)):
        for j in range(len(shifts[i])):
            if shifts[i][j] is None:
                shifts[i][j] = 0

    for k in range(len(shifts)):
        for i in range(0, 21):
            if shifts[k][i] < 60:
                shifts[k][i - 1] = int(shifts[k][i - 1]) + int(shifts[k][i])
                shifts[k][i] = 0
            else:
                shifts[k][i] = int(shifts[k][i]) - 30
                shifts[k][i - 1] = int(shifts[k][i - 1]) + 30

    for k in range(len(shifts)):
        for i in range(0, 21):
            if 30 > shifts[k][i] > 0 and shifts[k][i + 1] > 0:
                shifts[k][i + 1] = int(shifts[k][i + 1]) + int(shifts[k][i])
                shifts[k][i] = 0

    for row, lists in zip(matching_rows, shifts):
        for col, element in zip(range(13, 34), lists):
            worksheet.cell(row=row, column=col).value = element

    workbook.save(input_excel_path)
