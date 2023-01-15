# Author: Ozan Sahin

# Tasks to be done for standalone executable (in semester break):
# TODO: Delete print statements
# TODO: Integrate input validation with the UI

from importlib.util import find_spec
from subprocess import check_call, check_output
from sys import exit
import os
import sys


# This will not be needed when the script is converted to a standalone executable
def package_control(packages: list):
    for package in packages:
        if find_spec(package) is None:
            print(f"\n>>> Installing {package}...\n")
            check_call([sys.executable, '-m', 'pip', 'install', package,
                       '--disable-pip-version-check'])


package_control(packages=["pandas", "openpyxl", "numpy"])


# import modules after checking if they exist in the environment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import openpyxl
import pandas as pd


# Colors for Excel formatting
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

# separators for different operating systems are used to make it compatible with all of them
# directory separator = "/" for Linux and MacOS
# directory separator = "\\" for Windows

def file_path_handler():
    # validation added to fix VSCode issue (will be removed later)
    if str(os.getcwd()).split(os.sep)[-1] == "Code":
        directory = check_output(["python", f"{os.getcwd()}{os.sep}transformer_ui.py"])
    else:
        directory = check_output(["python", f"{os.getcwd()}{os.sep}Code{os.sep}transformer_ui.py"])
    directory = directory.decode("utf-8")
    directory = str(directory.strip())
    paths = {}

    # Create a dictionary from the output of the UI
    # The dictionary will be used to extract the file paths
    for i in directory.split(os.linesep):
        component = i.split("=")
        if len(component) > 1:
            paths[component[0]] = component[1]

    paths = dict(sorted(paths.items(), key=lambda item: item[0], reverse=True))

    # input validation for file paths 
    # TODO: Find an elegant way to do this and integrate it with the UI
    if "Source" in paths.keys() and "Output" in paths.keys():
        if paths["Source"] == "":
            print("!!> You have not selected an Excel file!")
            exit(0)
        elif paths["Output"] == "":
            print("!!> You have not selected an output destination!")
            exit(0)
    if "Source" in paths.keys() and not "Output" in paths.keys():
        if paths["Source"] == "":
            print("!!> You have not selected an Excel file and output destination!")
            exit(0)
        else:
            print("!!> You have not selected an output destination!")
            exit(0)
    if "Output" in paths.keys() and not "Source" in paths.keys():
        if paths["Output"] == "":
            print("!!> You have not selected an Excel file and output destination!")
            exit(0)
        else:
            print("!!> You have not selected an Excel file!")
            exit(0)
    if "Output" not in paths.keys() and "Source" not in paths.keys():
        print("!!> You have not selected an Excel file and output destination!")
        exit(0)

    # create output directory name from the source file path
    source_dir, output_dir = paths["Source"], paths["Output"]
    source_file_name = source_dir.split("/")[-1]
    output_dir = output_dir + os.sep + source_file_name.split(".")[0] + "_output.xlsx"
    

    try:
        source_file = pd.read_excel(source_dir)
    except FileNotFoundError:
        print("!!> File not found!")
        exit(0)

    # Validation will not be needed after the standalone executable
    if str(os.getcwd()).split(os.sep)[-1] == "Code":
        pipes_path = os.sep.join(str(os.getcwd()).split(os.sep)[:-1]) + \
            f"{os.sep}Data{os.sep}Cihazlar - Borular.xlsx"
        types_path = os.sep.join(str(os.getcwd()).split(os.sep)[:-1]) + \
            f"{os.sep}Data{os.sep}Borular - Tipler.xlsx"
    else:
        pipes_path = os.getcwd() + f"{os.sep}Data{os.sep}Cihazlar - Borular.xlsx"
        types_path = os.getcwd() + f"{os.sep}Data{os.sep}Borular - Tipler.xlsx"
    try:
        pipes = pd.read_excel(pipes_path)
        types = pd.read_excel(types_path)
    except FileNotFoundError:
        print("!!> File not found!")
        exit(0)

    return source_file, pipes, types, output_dir


# TODO: This will be not needed when the script is converted to a standalone executable
# Check if the Python version is 3.X.X (preferably, 3.11.X) if not raise an error
def python_version_control() -> None:
    if sys.version_info.major != 3:
        raise SystemError("!!> Python version must be 3.X.X (preferably, 3.11.X)")
    if sys.version_info.minor != 11:
        print("??> This script is written for Python 3.11.X, "
              "It may not work properly with other versions.\n"
              "??> Do you still want to continue? (Y/N)")
        user_input = input(">>> ")
        if user_input == 'Y':
            pass
        elif user_input == 'N':
            print(">>> Terminating...")
            exit(0)
        else:
            raise SystemError("!!> Invalid input!")


# General formatting such as column width, row height, and coloring
# Adds auto-filter to every column except the first one (index column)
def general_excel_formatter(file_path: str) -> None:
    wb = openpyxl.load_workbook(file_path)

    ws1 = wb["Sheet1"]
    ws1.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws1)

    for col in range(ws1.min_column, ws1.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws1, min=col, max=col, width=12)

    #  change the height of all rows
    for row in range(ws1.min_row, ws1.max_row + 1):
        ws1.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws1, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws1, min=3, max=3, width=18)
    dim_holder['D'] = ColumnDimension(ws1, min=4, max=4, width=18)
    dim_holder['E'] = ColumnDimension(ws1, min=5, max=5, width=18)

    # add filter
    ws1.auto_filter.ref = "A2:Z2"

    # highlight the version and date cells
    ws1['A1'].fill = redFill
    ws1['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1['B1'].fill = redFill
    ws1['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1.column_dimensions = dim_holder
    wb.save(file_path)


# similar to general_excel_formatter but for the pivotted sheet
def pivot_excel_formatter(file_path: str) -> None:
    wb = openpyxl.load_workbook(file_path)

    ws2 = wb["Sheet2"]

    ws2.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws2)

    for col in range(ws2.min_column, ws2.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws2, min=col, max=col, width=12)

    #  change the height of all rows
    for row in range(ws2.min_row, ws2.max_row + 1):
        ws2.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws2, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws2, min=3, max=3, width=18)

    # add filter
    ws2.auto_filter.ref = "A2:Z2"

    # highlight the version and date cells
    ws2['A1'].fill = redFill
    ws2['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2['B1'].fill = redFill
    ws2['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2.column_dimensions = dim_holder
    wb.save(file_path)


# Adds the version and date information to the excel file
def excel_version(file_path: str, file: pd.DataFrame) -> None:
    wb = openpyxl.load_workbook(file_path)

    version = file.iloc[[3, 4], 7]
    update_date = file.iloc[[3, 4], 8]

    version_value = version.iloc[0] + " - " + version.iloc[1]
    update_date_value = update_date.iloc[0] + ":  " + update_date.iloc[1]

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.cell(row=1, column=1).value = version_value
        ws.cell(row=1, column=2).value = str(update_date_value)
        ws.cell(row=2, column=1).value = "Hat"

        # center all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_path)


# checks if the excel file exists, if it does not then it creates it
def check_and_create_sheet(output_excel_file: str) -> None:

    print(">>>\n>>> Validating Excel Files...")
    try:
        if os.path.exists(output_excel_file):
            wb = openpyxl.load_workbook(output_excel_file)

            if "Sheet1" not in wb.sheetnames:
                wb.create_sheet("Sheet1")
            if "Sheet2" not in wb.sheetnames:
                wb.create_sheet("Sheet2")

            wb.save(output_excel_file)

        print(">>> Validation completed!")
    except PermissionError:
        print("!!> Permission denied!")
        print(">>> Terminating...")
        exit(1)
    except Exception as e:  # catch all other exceptions
        print(f"!!> {e}")
        print(">>> Terminating...")
        exit(1)


# writes the dataframes to the two sheets in the excel file (general-pivotted)
def write_to_excel(output_excel_file, main: pd.DataFrame, pivot: pd.DataFrame,
                   main_sheet_name="Sheet1", pivot_sheet_name="Sheet2") -> None:

    print(">>>\n>>> Conversion started...")

    try:
        with pd.ExcelWriter(output_excel_file, mode="w") as writer:
            main.to_excel(writer, main_sheet_name)
            pivot.to_excel(writer, pivot_sheet_name)

        print(">>> Conversion completed!")
    except PermissionError:
        print("!!> Conversion failed!")
        print("!!> Permission denied!")
        print(">>> Terminating...")
        exit(1)
    except Exception as e:  # catch all other exceptions
        print(f"!!> {e}")
        print(">>> Terminating...")
        exit(1)
