import os
import sys
import openpyxl
import pandas as pd
from tkinter import messagebox, filedialog
from customtkinter import CTk, CTkFrame, CTkLabel, CTkButton, CTkFont, CTkOptionMenu, CTkEntry, set_appearance_mode, \
    set_default_color_theme, CTkToplevel, CTkTabview
from rules import first_rule, third_rule


class App(CTk):
    set_appearance_mode("Dark")
    set_default_color_theme("dark-blue")

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)

        self.current_directory = os.path.dirname(os.getcwd() + f"{os.sep}Data{os.sep}Source{os.sep}")
        self.first_file_name, self.second_file_name, self.directory_name = "", "", ""

        self.shift_window = None
        self.close = None
        self.window_configuration()
        self.grid_configuration()

        # create sidebar frame with widgets
        self.sidebar_frame = CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=3, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(6, weight=1)
        self.logo_label = CTkLabel(self.sidebar_frame, text="Excel Converter",
                                   font=CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.shifts_button, self.next_week_button, self.current_week_button, self.output_destination_button, \
            self.transform_button = self.create_buttons()
        self.next_week_path, self.current_week_path, self.output_destination_path = self.create_labels()
        self.place_components()

        self.appearance_mode_label = CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=6, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_option_menu = CTkOptionMenu(self.sidebar_frame,
                                                         values=["Dark", "Light", "System"],
                                                         command=self.change_appearance_mode_event)
        self.appearance_mode_option_menu.grid(row=7, column=0, padx=20, pady=(10, 20))
        self.create_labels()

        self.button_validation()

    def button_validation(self) -> None:
        """
        Sets the default values of the widgets and controls the button activation.
        """
        self.next_week_path.configure(justify="center")
        self.current_week_path.configure(justify="center")
        self.output_destination_path.configure(justify="center")
        self.appearance_mode_option_menu.set("Dark")

        # check the current values and activate the buttons accordingly
        if self.next_week_path.get() == "" or self.next_week_path.get() == "Next Week":
            self.current_week_button.configure(state="disabled")
        else:
            self.current_week_button.configure(state="normal")
        if self.current_week_path.get() == "" or self.current_week_path.get() == "Current Week":
            self.output_destination_button.configure(state="disabled")
        else:
            self.output_destination_button.configure(state="normal")
        if self.output_destination_path.get() == "" or self.output_destination_path.get() == "Output Destination" \
                or self.next_week_path.get() == "" or self.next_week_path.get() == "Next Week" or \
                self.current_week_path.get() == "" or self.current_week_path.get() == "Current Week":
            self.shifts_button.configure(state="disabled")
        else:
            self.shifts_button.configure(state="normal")
        if self.output_destination_path.get() == "" or self.current_week_path.get() == "" or \
                self.next_week_path.get() == "" or self.output_destination_path.get() == "Output Destination" or \
                self.current_week_path.get() == "Current Week" or self.next_week_path.get() == "Next Week" \
                and self.shifts_button["state"] != "disabled":
            self.transform_button.configure(state="disabled")
        else:
            self.transform_button.configure(state="normal")

    def on_close(self) -> None:
        """
        Asks the user whether they want to close the program.
        """
        self.close = messagebox.askokcancel("Close", "Would you like to close the program?",
                                            icon="warning", parent=self)

        if self.close:
            self.destroy()
            sys.exit(0)

    @staticmethod
    def show_error(message) -> None:
        """
        Shows an error message.

        Args:
            message: The error message that will be displayed. (e.g. "You have not selected the first Excel file!")
        """
        message_window = CTk()
        message_window.withdraw()
        messagebox.showerror("Error", message, icon="error")
        message_window.destroy()
        sys.exit(0)

    @staticmethod
    def default_font() -> tuple[str, int]:
        """
        Checks whether the operating system is Windows or macOS and selects the default font and font size accordingly.

        Returns:
            A tuple with the default font and font size.
        """
        if sys.platform == "win32":
            return "Arial", int(10)
        elif sys.platform == "darwin":
            return "Courier", int(13)

    def transform(self):
        first_rule(input_excel_path=self.next_week_path)
        third_rule(input_excel_path=self.next_week_path)
        self.destroy()

    def create_buttons(self) -> tuple[CTkButton, CTkButton, CTkButton, CTkButton, CTkButton]:
        """
        Creates the buttons in the UI. Input1 and input2 are the buttons that are used to select the Excel files for
        the next week's and current week's production plan. Output is the button that is used to select the output
        directory for the transformed Excel file. Transform is the button that is used to start the transformation.

        Returns:
            A tuple with the next week, current week, output destination, and transform buttons.
        """
        shifts_button = CTkButton(self.sidebar_frame, command=self.get_shifts,
                                  text="Rules")
        next_week_button = CTkButton(self.sidebar_frame, command=self.browse_first_input_file,
                                     text="Next Week")
        current_week_button = CTkButton(self.sidebar_frame, command=self.browse_second_input_file,
                                        text="Current Week")
        output_destination_button = CTkButton(self.sidebar_frame, command=self.browse_output_directory,
                                              text="Output Destination")
        transform_button = CTkButton(self.sidebar_frame, command=self.destroy,
                                     text="Transform")
        return shifts_button, next_week_button, current_week_button, output_destination_button, transform_button

    def create_labels(self) -> tuple[CTkEntry, CTkEntry, CTkEntry]:
        """
        Creates the labels in the UI. Input1 and input2 are the text boxes where the user can see the selected file
        path. Output is the text box where the user can see the selected output directory. Progress is the text bar
        that is used to show the progress of the transformation.

        Returns:
            A tuple with the next week, current week, output destination, and progress labels.
        """
        next_week_path = CTkEntry(self, placeholder_text="Next Week")
        next_week_path.configure(state="disabled")
        next_week_path.xview_scroll(1, "units")
        current_week_path = CTkEntry(self, placeholder_text="Current Week")
        current_week_path.configure(state="disabled")
        output_destination_path = CTkEntry(self, placeholder_text="Output Destination")
        output_destination_path.configure(state="disabled")
        return next_week_path, current_week_path, output_destination_path

    def place_components(self) -> None:
        """
        Places the components such as buttons and labels in the UI.
        """
        self.next_week_button.grid(row=1, column=0, padx=20, pady=10)
        self.current_week_button.grid(row=2, column=0, padx=20, pady=10)
        self.output_destination_button.grid(row=3, column=0, padx=20, pady=10)
        self.shifts_button.grid(row=4, column=0, padx=20, pady=10)
        self.transform_button.grid(row=5, column=0, padx=20, pady=10)

        self.next_week_path.grid(row=0, column=1, padx=(20, 20), pady=(20, 0), sticky="nsew")
        self.current_week_path.grid(row=1, column=1, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.output_destination_path.grid(row=2, column=1, padx=(20, 20), pady=(0, 20), sticky="nsew")

    def grid_configuration(self) -> None:
        """
        Configures the row and column weights of the grid. The grid will be placed in the window to align the buttons
        and the labels of the UI.
        """
        # center the buttons and window
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=0)
        self.grid_columnconfigure(3, weight=0)

        # make the windows expandable
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=0)

    def window_configuration(self) -> None:
        """
        Sets the title of the window, makes the window non-resizable, and makes the window close when the user clicks.
        Also sets the dimensions of the window and places it in the center of the screen.
        """
        self.title("Excel Converter")

        # make the window non-resizable
        self.resizable(False, False)

        # make the window close when the user clicks the X button
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # set the dimensions of the window
        w = 750  # width for the Tk root
        h = 393  # height for the Tk root

        # get screen width and height
        ws = self.winfo_screenwidth()  # width of the screen
        hs = self.winfo_screenheight()  # height of the screen

        # calculate x and y coordinates for the Tk root window
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)

        # set the dimensions of the screen where the window will be displayed
        self.geometry('%dx%d+%d+%d' % (w, h, x, y))

    def get_shifts(self):
        if self.shift_window is None or not self.shift_window.winfo_exists():
            self.shift_window = ShiftWindow(self, next_week_excel_path=self.next_week_path.get(),
                                            current_week_excel_path=self.current_week_path.get())
            self.shift_window.focus_force()
            self.shift_window.grab_set()

    def browse_first_input_file(self) -> None:
        """
        Opens a file dialog box to select the next week's production plan Excel file. The file dialog box will only
        show Excel files.
        """
        self.first_file_name = filedialog.askopenfilename(initialdir=self.current_directory,
                                                          title="Select a File (Next Week)",
                                                          filetypes=(("Excel Files", "*.xlsx"),
                                                                     ("Excel Macro Files", "*.xlsm"),))
        print(f"Source1={self.first_file_name}")
        self.next_week_path.configure(state="normal")
        if self.next_week_path.get() != "":
            self.next_week_path.delete(0, "end")
        if self.first_file_name != "":
            self.next_week_path.insert(index=0, string=self.first_file_name)
        else:
            self.next_week_path.insert(index=0, string="Next Week")
        self.next_week_path.configure(state="disabled")
        self.button_validation()

    def browse_second_input_file(self) -> None:
        """
        Opens a file dialog box to select the current week's production plan Excel file. The file dialog box will
        only show Excel files.
        """
        self.second_file_name = filedialog.askopenfilename(initialdir=self.current_directory,
                                                           title="Select a File (Current Week)",
                                                           filetypes=(("Excel Files", "*.xlsx"),
                                                                      ("Excel Macro Files", "*.xlsm"),))
        print(f"Source2={self.second_file_name}")
        self.current_week_path.configure(state="normal")
        if self.current_week_path.get() != "":
            self.current_week_path.delete(0, "end")
        if self.second_file_name != "":
            self.current_week_path.insert(index=0, string=self.second_file_name)
        else:
            self.current_week_path.insert(index=0, string="Current Week")
        self.current_week_path.configure(state="disabled")
        self.button_validation()

    def browse_output_directory(self) -> None:
        """
        Opens a directory dialog box to select the output directory where the transformed Excel file will be saved.
        """
        self.directory_name = filedialog.askdirectory(initialdir=self.current_directory,
                                                      title="Select a Directory")
        print(f"Output={self.directory_name}")
        self.output_destination_path.configure(state="normal")
        if self.output_destination_path.get() != "":
            self.output_destination_path.delete(0, "end")
        if self.directory_name != "":
            self.output_destination_path.insert(index=0, string=self.directory_name)
        else:
            self.output_destination_path.insert(index=0, string="Output Destination")
        self.output_destination_path.configure(state="disabled")
        self.button_validation()

    @staticmethod
    def change_appearance_mode_event(new_appearance_mode: str) -> None:
        """
        Changes the appearance mode of the UI.

        Args:
            new_appearance_mode: The new appearance mode of the UI. (light or dark)
        """
        set_appearance_mode(new_appearance_mode)


class ShiftWindow(CTkToplevel):

    def __init__(self, *args, next_week_excel_path, current_week_excel_path, **kwargs):
        super().__init__(*args, **kwargs)

        self.dates = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.score_labels = []
        self.score_entries = []
        self.next_week_excel_path = next_week_excel_path
        self.current_week_excel_path = current_week_excel_path
        self.window_configuration()
        self.create_grid()

    def window_configuration(self) -> None:
        """
        Sets the title of the window, makes the window non-resizable, and makes the window close when the user clicks.
        Also sets the dimensions of the window and places it in the center of the screen.
        """
        self.title("Shift Selection")

        # make the window non-resizable (for macOS and Linux)
        if sys.platform == "darwin":
            self.resizable(False, False)

        # set the dimensions of the window
        w = 565  # width for the Tk root
        h = 420  # height for the Tk root

        # get screen width and height
        ws = self.winfo_screenwidth()  # width of the screen
        hs = self.winfo_screenheight()  # height of the screen

        # calculate x and y coordinates for the Tk root window
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)

        # set the dimensions of the screen where the window will be displayed
        self.geometry('%dx%d+%d+%d' % (w, h, x, y))

    def save_data(self):
        data = {'Gün': [date for date in self.dates]}
        df = pd.DataFrame(data)
        for shift in range(3):
            data[f'Shift {shift + 1}'] = [self.score_entries[self.dates.index(date) * 3 + shift].get() for date in
                                          self.dates]

        workbook = openpyxl.load_workbook(self.next_week_excel_path)
        worksheet = workbook['GZT-GWT']

        last_row = worksheet.max_row
        matching_rows = [i for i in range(1, last_row + 1) if
                         worksheet.cell(row=i, column=8).value and str(
                             worksheet.cell(row=i, column=8).value).startswith(
                             "7")]

        shifts = []

        for row in matching_rows:
            row_vardiya = []
            for col in range(13, 34):
                row_vardiya.append(worksheet.cell(row=row, column=col).value)
            shifts.append(row_vardiya)

        for i in range(len(shifts)):
            for j in range(len(shifts[i])):
                if shifts[i][j] is None:
                    shifts[i][j] = 0

        shift_list = []
        for i in range(len(df)):
            for j in range(len(df.columns)):
                if df.iloc[i, j] == "Var":
                    shift_list.append((i * 3) + j - 1)

        for k in range(len(shifts)):
            for i in shift_list:
                shifts[k][i - 1] = int(shifts[k][i]) + int(shifts[k][i - 1])
                shifts[k][i] = 0

        for row, lists in zip(matching_rows, shifts):
            for col, element in zip(range(13, 34), lists):
                worksheet.cell(row=row, column=col).value = element

        # DISABLED FOR DEBUGGING PURPOSES!
        # workbook.save(self.next_week_excel_path)

        self.destroy()

    def create_grid(self):
        current_week_tab_name = "Current Week"
        next_week_tab_name = "Next Week"

        current_week_identifier = self.current_week_excel_path.split("/")[-1].split(".")[0]
        next_week_identifier = self.next_week_excel_path.split("/")[-1].split(".")[0]

        tabview = CTkTabview(self)
        tabview.pack(fill="both", expand=True, pady=(0, 18))

        tabview.add(current_week_tab_name)
        tabview.add(next_week_tab_name)

        tab_names = [current_week_tab_name, next_week_tab_name]
        week_identifiers = [current_week_identifier, next_week_identifier]

        for index, tab_name in enumerate(tab_names):
            week_identifier = CTkLabel(tabview.tab(tab_name), text=week_identifiers[index])
            week_identifier.grid(row=0, column=0)

            header_labels = [CTkLabel(tabview.tab(tab_name), text=f"Shift {i + 1}") for i in range(3)]
            for i, header in enumerate(header_labels):
                header.grid(row=0, column=i + 1, sticky="nsew", padx=5, pady=5)

            for date in self.dates:
                date_label = CTkLabel(tabview.tab(tab_name), text=date)
                date_label.grid(row=self.dates.index(date) + 1, column=0, sticky="nsew", padx=15, pady=5)

                for shift in range(3):
                    score_entry = CTkOptionMenu(tabview.tab(tab_name), values=["Yok", "Var"], anchor="center")
                    score_entry.grid(row=self.dates.index(date) + 1, column=shift + 1, padx=5, pady=5, sticky="nsew")

                    self.score_entries.append(score_entry)
                    self.score_labels.append(date)

            save_button = CTkButton(tabview.tab(tab_name), text="Save", command=self.save_data)
            save_button.grid(row=len(self.dates) + 2, column=3, padx=5, pady=10)


if __name__ == "__main__":
    app = App()
    app.mainloop()
