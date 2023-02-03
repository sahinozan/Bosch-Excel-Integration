import pandas as pd
import openpyxl


#Bu haftanın Cumartesi ve Pazar günleri --------------------------------------
df1 = pd.read_excel("KW47_V05.xlsx")

kolon1 = df1.iloc[:, 7]

cihaz_ttnr1 = kolon1[kolon1.astype(str).str.startswith(('7', '8'))]
hat1 = df1.iloc[cihaz_ttnr1.index, 0]
aile1 = df1.iloc[cihaz_ttnr1.index, 8]
üretim1 = df1.iloc[cihaz_ttnr1.index, 27:33]

data1 = {
        'Hat': hat1, 
        'Cihaz TTNR': cihaz_ttnr1, 
        'Aile': aile1,
        }

df1 = pd.DataFrame(data1)
veri1 = pd.concat([df1, üretim1], axis = 1)


#Gelecek haftanın Pazartesi, Salı, Çarşamba, Perşembe ve Cuma günleri --------
df2 = pd.read_excel("KW48_Taslak_Plan.xlsx")

kolon2 = df2.iloc[:, 7]

cihaz_ttnr2 = kolon2[kolon2.astype(str).str.startswith(('7', '8'))]
hat2 = df2.iloc[cihaz_ttnr2.index, 0]
aile2 = df2.iloc[cihaz_ttnr2.index, 8]
üretim2 = df2.iloc[cihaz_ttnr2.index, 12:27]

data2 = {
        'Hat': hat2, 
        'Cihaz TTNR': cihaz_ttnr2, 
        'Aile': aile2,
        }

df2 = pd.DataFrame(data2)
veri2 = pd.concat([df2, üretim2], axis = 1)

#Tabloları birleştirme -------------------------------------------------------
veri3 = pd.concat([veri1, veri2])

veri3.sort_values(by=veri3.columns[0], axis=0, ascending=True, inplace=True)

veri3.to_excel('output.xlsx', index=False)

#Sütun isimlerini değiştir: 1-2-3 --------------------------------------------
wb = openpyxl.load_workbook('output.xlsx')
ws = wb['Sheet1']

counter = 1
for col in range(4, 25):
    cell = ws.cell(row=1, column=col)
    cell.value = counter
    counter = counter + 1
    if counter > 3:
        counter = 1

#Günler, Merge ve Ortalama, filtre -------------------------------------------
ws.insert_rows(1)
days = ["Cumartesi", "Pazar", "Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

i = 4
for i, day in enumerate(days):
    ws.cell(row=1, column=i*3+4).value = day

for i in range(4, 25, 3):
    ws.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+2)
    
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

wb.save("output.xlsx")