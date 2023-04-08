# Author: Ozan Şahin

from __future__ import annotations
import datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from importlib.util import find_spec
from subprocess import check_call, check_output
import os
import sys
from custom_ui import App


def package_control(packages: list) -> None:
    """
    Check if the required packages are installed and install them if not.
    For each non-installed package, pip install command will be executed.

    Args:
        packages: List of packages to be checked
    """
    for package in packages:
        if find_spec(package) is None:
            check_call([sys.executable, '-m', 'pip', 'install', package,
                        '--disable-pip-version-check'])


package_control(packages=["pandas", "openpyxl", "numpy", "customtkinter"])

# Colors for Excel formatting
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')


def file_path_handler() -> \
        tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str, str]:
    """
    Get the file paths from the UI, convert the files into dataframes.

    Returns:
        A tuple which contains the following:
        current_source_file: A dataframe of the current week's production plan
        past_source_file: A dataframe of the next week's production plan
        pipes: A dataframe that contains the amount of pipes for each device
        types: A dataframe of the pipe types (e.g. hydraulic, spare, etc.)
        output_dir: A string of the output directory
    """
    # Run the UI and get the file paths
    if str(os.getcwd()).split(os.sep)[-1] == "Code":
        directory = check_output(["python", f"{os.getcwd()}{os.sep}custom_ui.py"])
    else:
        directory = check_output(["python", f"{os.getcwd()}{os.sep}Code{os.sep}custom_ui.py"])
    directory = directory.decode("utf-8")
    directory = str(directory.strip())
    paths = {}

    # Create a dictionary from the output of the UI
    # The dictionary will be used to extract the file paths
    for i in directory.split(os.linesep):
        component = i.split("=")
        if len(component) > 1:
            paths[component[0]] = component[1]

    sorted_paths = sorted(paths.items(), key=lambda item: item[0], reverse=True)
    paths = {key: path for key, path in sorted_paths}

    # input validation for file paths
    path_validation(paths)

    # create output directory name from the source file path
    source_dir, output_dir = paths["Source"], paths["Output"]
    source_file_name = source_dir.split("/")[-1]
    output_dir = output_dir + os.sep + source_file_name.split(".")[0] + "_output.xlsx"

    try:
        source_file = pd.read_excel(source_dir)
    except FileNotFoundError:
        App.show_error("File not found!")
        sys.exit(0)

    # Validation will not be needed after the standalone executable
    if str(os.getcwd()).split(os.sep)[-1] == "Code":
        master_path = os.sep.join(str(os.getcwd()).split(os.sep)[:-1]) + \
                      f"{os.sep}Data{os.sep}Master Data.xlsx"
    else:
        master_path = os.getcwd() + f"{os.sep}Data{os.sep}Master Data.xlsx"
    try:
        pipes = pd.read_excel(master_path, sheet_name="Cihaz - Boru - Miktar")
        types = pd.read_excel(master_path, sheet_name="Boru - Tip")
    except FileNotFoundError:
        App.show_error("File not found!")
        sys.exit(0)

    return source_file, pipes, types, output_dir, source_dir


def total_quantity_per_pipe(output_excel_file_path) -> None:
    """
    Sum the quantity of each shift for each pipe and add it to the add to an extra column.

    Args:
        output_excel_file_path: The path of the output Excel file
    """
    wb = openpyxl.load_workbook(output_excel_file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        if sheet == "Genel":
            ws.insert_cols(27, 1)
            ws["AA1"] = "Toplam"
            ws["AA1"].border = Border(left=Side(border_style="thin", color='FF000000'),
                                      right=Side(border_style="thin", color='FF000000'),
                                      top=Side(border_style="thin", color='FF000000'),
                                      bottom=Side(border_style="thin", color='FF000000'))
            ws["AA2"] = "X"
            ws["AA2"].border = Border(left=Side(border_style="thin", color='FF000000'),
                                      right=Side(border_style="thin", color='FF000000'),
                                      top=Side(border_style="thin", color='FF000000'),
                                      bottom=Side(border_style="thin", color='FF000000'))
            for i in range(3, ws.max_row):
                ws[f"AA{i + 1}"] = f"=SUM(F{i}:Z{i})"
        elif sheet == "Pivot":
            ws.insert_cols(25, 1)
            ws["Y1"] = "Toplam"
            ws["Y1"].border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'))
            ws["Y2"] = "X"
            ws["Y2"].border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'))
            for i in range(3, ws.max_row):
                ws[f"Y{i + 1}"] = f"=SUM(F{i}:X{i})"

    wb.save(output_excel_file_path)


def path_validation(paths: dict) -> None:
    """
    Checks whether the user has selected the Excel files and the output destination.

    Args:
        paths: A dictionary containing the paths of the Excel files and the output destination.
    """
    if "Source" not in paths.keys() or "Output" not in paths.keys():
        sys.exit(1)


def general_excel_converter(raw_df: pd.DataFrame, pipes: pd.DataFrame, types: pd.DataFrame) -> \
        pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Filters the rows with missing values and eliminates the unnecessary columns in the main dataframe.
    Then, it merges the main dataframe with the pipes and types dataframes.
    As a result, the final dataframe contains the type and the amount of pipes for each device.

    Args:
        raw_df: Main dataframe that contains the raw data (e.g. the current week's production plan)
        pipes: A dataframe that contains the amount of pipes for each device
        types: A dataframe of the pipe types (e.g. hydraulic, spare, etc.)

    Returns:
        A dataframe with multi-level columns
    """
    # get the shift dates and format them (e.g. 27 Dec 2022)
    date_start_index = str(raw_df.columns[raw_df.isin(['Pazartesi']).any()][0]).split(' ')[1]

    shift_date = raw_df.iloc[4:6, int(date_start_index): 31: 3].copy()
    shift_date.iloc[0, :] = shift_date.iloc[0, :].apply(lambda x: x.strftime("%d %b %Y"))
    shift_date = shift_date.apply(lambda x: f"{x.iloc[1]} - {x.iloc[0]}", axis=0)
    shift_dates = list(shift_date)

    # TTNr, Hat, Cihaz Aile, and work days columns
    indices = raw_df.iloc[:, [0, 7, 8, 11]].reset_index()
    work_days = raw_df.iloc[:, 12: 33].reset_index()
    sheet = pd.concat([indices, work_days], axis=1).iloc[2:, :]

    # drop the rows with NaN values in the TTNr column
    sheet = sheet[sheet.iloc[:, 1].notna() & sheet.iloc[:, 2].notna()]
    sheet.iloc[:, 2] = sheet.iloc[:, 2].astype(str)

    sheet = sheet[sheet.loc[:, "MOE1 Üretim Sıralaması"].apply(
        lambda x: (len([j for j in x if j.isnumeric()]) > 1 and len(x) >= 3))]

    # get only the numeric values
    sheet = sheet[sheet.iloc[:, 6].apply(lambda x: (type(x) != datetime.datetime) and (type(x) != str))]
    sheet = sheet[sheet.iloc[:, 3].notna()]

    # drop the index column (unnecessary)
    sheet.drop('index', axis=1, inplace=True)
    sheet.reset_index(drop=True, inplace=True)

    # create the final dataframe indices
    initial_indices = ["Hat", "Cihaz TTNr", "Cihaz Aile", "Toplam Adet"]
    shifts = [1, 2, 3]
    final_indices = [" - ".join([i, str(j)]) for i in shift_dates for j in shifts]
    initial_indices.extend(final_indices)

    sheet = sheet.set_axis(initial_indices, axis=1)

    # convert to string to for the merge operation
    sheet["Cihaz TTNr"] = sheet["Cihaz TTNr"].astype(str)
    pipes["Cihaz"] = pipes["Cihaz"].astype(str)

    # merge the pipes and devices dataframes
    sheet = sheet.merge(pipes, left_on="Cihaz TTNr", right_on="Cihaz", how="inner")
    sheet = sheet.merge(types, on="Boru", how="left")

    # drop the unnecessary columns
    sheet.drop("Cihaz", axis=1, inplace=True)
    sheet.insert(3, 'Boru', sheet.pop('Boru'))

    indices = ["Hat", "Cihaz TTNr", "Cihaz Aile", "Boru TTNr"]

    # create the final dataframe with multi-level columns (same format with the initial Excel)
    df = pd.DataFrame(columns=pd.MultiIndex.from_product([shift_dates, shifts]),
                      index=range(sheet.shape[0]))
    df = pd.concat([pd.DataFrame(columns=pd.MultiIndex.from_product(
        [indices, ["" for _ in range(len(indices))]])), df], axis=1)

    # swap levels according to the initial Excel and drop the duplicated columns
    dates_df = df.iloc[:, 16:]
    initial_df = df.swaplevel(axis=1, i=0, j=1).iloc[:, :16]
    initial_df = initial_df.loc[:, ~initial_df.columns.duplicated()]
    df = pd.concat([initial_df, dates_df], axis=1)
    sheet.iloc[:, 4:-1] = sheet.iloc[:, 4:-1].apply(lambda x: x * sheet.iloc[:, -2], axis=0)

    # drop the unnecessary columns
    sheet.drop("Miktar", axis=1, inplace=True)
    sheet.drop("Toplam Adet", axis=1, inplace=True)

    # add the Tip column to the dataframe for the merge operation
    df[df.columns] = sheet[sheet.columns[:-1]]

    # add the Tip column to the df
    df["Tip"] = sheet["Tip"]

    # merge the dataframe with the types dataframe (hydraulic, spare, etc.)
    type_df = df.swaplevel(axis=1, i=0, j=1).iloc[:, -1]
    df = pd.concat([df.iloc[:, :-1], type_df], axis=1)
    df.insert(4, ('', 'Tip'), df.pop(('', 'Tip')))

    # dropped the index column name (will be filled later with openpyxl for better visuals)
    df = df.set_index(("", "Hat")).rename_axis(axis=0)
    df.index = df.index.map(lambda x: "Hat 1" if '7' in x else f"Hat {x.split(' ')[1]}")

    # convert work days columns to numeric values
    df[df.columns[4]] = df[df.columns[4]].apply(pd.to_numeric, errors='coerce')

    return df


def excel_pivoting(df_initial: pd.DataFrame, types: pd.DataFrame) -> pd.DataFrame:
    """
    Does pivoting operation for the master dataframe. The pivoting combines the same pipes in the same shift.

    Args:
        df_initial: The master dataframe that contains the type and the amount of pipes for each device
        types: A dataframe that contains the types of the pipes (e.g. hydraulic, spare, etc.)

    Returns:
        A dataframe that contains the pivoted values
    """
    # create the dataframe for the Excel pivoting (multi-level columns)
    df_pivoted = df_initial.sort_index(key=lambda x: (x.to_series().str[4:].astype("int64")))
    df_pivoted = df_pivoted.drop(columns=[('', 'Cihaz TTNr'), ('', 'Cihaz Aile'), ('', 'Tip')])

    # sum the values for the same pipe and shift (pivoting)
    df_pivoted.index = df_pivoted.index.str.split(' ').str[1]
    df_pivoted = df_pivoted.groupby([df_pivoted.index, ("", "Boru TTNr")]).sum(numeric_only=False).sort_index(
        ascending=False)
    df_pivoted = df_pivoted.reset_index(level=1, drop=False)
    df_pivoted.index = df_pivoted.index.map(lambda x: x == "Hat 1" if x == 7 else f'Hat {x}')
    df_pivoted["Tip"] = df_pivoted.loc[:, ("", "Boru TTNr")].map(types.set_index("Boru")["Tip"])

    # swap the levels of the columns to match the initial Excel
    swapped_types = df_pivoted.iloc[:, -1].to_frame().swaplevel(axis=1, i=0, j=1).iloc[:, 0].to_frame()
    df_pivoted.insert(1, ('', 'Tip'), swapped_types)
    df_pivoted.drop(("Tip", ""), axis=1, inplace=True)

    # convert nan values to 0 to prevent errors in Excel
    df_pivoted.iloc[:, 2:] = df_pivoted.iloc[:, 2:].applymap(lambda x: np.nan if x == 0 else x)

    return df_pivoted


def excel_format_validate(list_of_dfs: list[pd.DataFrame]) -> None:
    """
    Checks the format of the given Excel file. If the format is not correct, the program will be terminated.

    Args:
        list_of_dfs: A list of dataframes of the Excel files
    """
    for i in list_of_dfs:
        if len(i.columns[i.isin(['Pazartesi']).any()]) == 0:
            App.show_error("Format of the first Excel file is not desired. Use an appropriate formatted Excel file.")
            sys.exit(0)
    pass


def general_excel_formatter(file_path: str, sheet_name) -> None:
    """
    Does general formatting such as column width, row height, and coloring.
    Adds auto-filter to every column except the index column.

    Args:
        file_path: The path of the Excel file
        sheet_name: The name of the sheet that will be formatted
    """
    wb = openpyxl.load_workbook(file_path)

    ws1 = wb[sheet_name]

    ws1.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws1)

    for col in range(ws1.min_column, ws1.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws1, min=col, max=col, width=12)

    # change the height of all rows
    for row in range(ws1.min_row, ws1.max_row + 1):
        ws1.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws1, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws1, min=3, max=3, width=18)
    dim_holder['D'] = ColumnDimension(ws1, min=4, max=4, width=18)
    dim_holder['E'] = ColumnDimension(ws1, min=5, max=5, width=18)

    # add filter
    ws1.auto_filter.ref = "A2:AA2"

    # highlight the version and date cells
    ws1['A1'].fill = redFill
    ws1['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1['B1'].fill = redFill
    ws1['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1['AA1'].fill = redFill
    ws1['AA1'].font = Font(color="FFFFFF", bold=True, size=11)
    ws1['AA2'].font = Font(color="000000", bold=True, size=11)

    for row in range(ws1.min_row, ws1.max_row + 1):
        for col in range(ws1.min_column, ws1.max_column + 1):
            if ws1.cell(row, col).value == 0:
                ws1.cell(row, col).value = np.nan

    ws1.column_dimensions = dim_holder
    wb.save(file_path)


def pivot_excel_formatter(file_path: str) -> None:
    """
    Does general formatting such as column width, row height, and coloring.
    Adds auto-filter to every column except the index column.

    Args:
        file_path: The path of the Excel file that will be formatted
    """
    wb = openpyxl.load_workbook(file_path)

    ws2 = wb["Pivot"]

    ws2.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws2)

    for col in range(ws2.min_column, ws2.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws2, min=col, max=col, width=12)

    # change the height of all rows
    for row in range(ws2.min_row, ws2.max_row + 1):
        ws2.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws2, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws2, min=3, max=3, width=18)

    # add filter
    ws2.auto_filter.ref = "A2:Y2"

    # highlight the version and date cells
    ws2['A1'].fill = redFill
    ws2['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2['B1'].fill = redFill
    ws2['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2['Y1'].fill = redFill
    ws2['Y1'].font = Font(color="FFFFFF", bold=True, size=11)
    ws2['Y2'].font = Font(color="000000", bold=True, size=11)

    ws2.column_dimensions = dim_holder
    wb.save(file_path)


def excel_version(file_path: str, file: pd.DataFrame) -> None:
    """
    Adds the version and date information to the Excel file.

    Args:
        file_path: The path of the Excel file that will be formatted
        file: The dataframe that contains the version and date information
    """
    wb = openpyxl.load_workbook(file_path)

    version = file.iloc[[3, 4], 7]
    update_date = file.iloc[[3, 4], 8]

    version_value = version.iloc[0] + " - " + version.iloc[1]
    update_date_value = update_date.iloc[0] + ":  " + update_date.iloc[1]

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.cell(row=1, column=1).value = version_value
        ws.cell(row=1, column=2).value = str(update_date_value)
        ws.cell(row=2, column=1).value = "Hat"

        # center all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_path)


def check_and_create_sheet(output_excel_file: str) -> None:
    """
    Checks if the Excel file exists, if it does not then it creates it.

    Args:
        output_excel_file: The path of the formatted Excel file
    """
    try:
        if os.path.exists(output_excel_file):
            wb = openpyxl.load_workbook(output_excel_file)

            if "Genel" not in wb.sheetnames:
                wb.create_sheet("Genel")
            if "Pivot" not in wb.sheetnames:
                wb.create_sheet("Pivot")

            wb.save(output_excel_file)
    except PermissionError:
        App.show_error("Permission Error!")
        sys.exit(0)
    except Exception as e:  # catch all other exceptions
        App.show_error(f"{e}!")
        sys.exit(0)


def write_to_excel(output_excel_file, main: pd.DataFrame, pivot: pd.DataFrame,
                   main_sheet_name="Genel", pivot_sheet_name="Pivot") -> None:
    """
    Writes the dataframes to the three separate sheets in the Excel file.

    Args:
        output_excel_file: The path of the formatted Excel file
        main: The dataframe that will be written to the general sheet (non-pivoted & formatted)
        pivot: The dataframe that will be written to the pivoted sheet (pivoted & formatted)
        main_sheet_name: The name of the general sheet
        pivot_sheet_name: The name of the pivoted sheet
    """
    try:
        with pd.ExcelWriter(output_excel_file, mode="w") as writer:
            main.to_excel(writer, main_sheet_name)
            pivot.to_excel(writer, pivot_sheet_name)
    except PermissionError:
        App.show_error("Conversion Failed!")
        App.show_error("Permission Error!")
        sys.exit(0)
    except Exception as e:  # catch all other exceptions
        App.show_error(f"{e}!")
        sys.exit(0)
