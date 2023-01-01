from __future__ import annotations
from importlib.util import find_spec
from sys import exit
import warnings
import subprocess
import datetime
import sys
import os

warnings.filterwarnings("ignore")

# check compatibility with Python 3.11
if sys.version_info.major != 3:
    raise SystemError("!!> Python version must be 3.X.X (preferably, 3.11.X)")
if sys.version_info.minor != 11:
    print("??> This script is written for Python 3.11.X, "
          "It may not work properly with other versions.\n"
          "??> Do you still want to continue? (Y/N)")
    user_input = input(">>> ")
    if user_input == 'Y':
        pass
    elif user_input == 'N':
        print(">>> Terminating...")
        exit(0)
    else:
        raise SystemError("!!> Invalid input!")

#  check if required packages are installed
if find_spec('pandas') is None:
    print("\n>>> Installing pandas...\n")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'pandas', '--disable-pip-version-check'])
if find_spec("openpyxl") is None:
    print("\n>>> Installing openpyxl...\n")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--disable-pip-version-check'])
if find_spec("numpy") is None:
    print("\n>>> Installing numpy...\n")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'numpy', '--disable-pip-version-check'])

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter



# read data source files
try:
    excel_file = input("\n>>> Enter the Excel file name: ")
    file = pd.read_excel(f'../Data/Source/{excel_file}.xlsx')
    pipes = pd.read_excel('../Data/Cihazlar - Borular.xlsx')
    types = pd.read_excel('../Data/Borular - Tipler.xlsx')
except FileNotFoundError:
    print("File not found!")
    exit(1)

#  get the date index range
date_start_index = file.columns[file.isin(['Pazartesi']).any()][0].split(' ')[1]

shift_date = file.iloc[4:6, int(date_start_index): 31: 3].copy()
shift_date.iloc[0, :] = shift_date.iloc[0, :].apply(lambda x: x.strftime("%d %b %Y"))
shift_date = shift_date.apply(lambda x: f"{x.iloc[1]} - {x.iloc[0]}", axis=0)
shift_dates = list(shift_date)

version = file.iloc[[3, 4], 7]
update_date = file.iloc[[3, 4], 8]

indices = file.iloc[:, [0, 7, 8, 11]].reset_index()
work_days = file.iloc[:, 12: 33].reset_index()
sheet = pd.concat([indices, work_days], axis=1).iloc[2:, :]

sheet = sheet[sheet.iloc[:, 1].notna() & sheet.iloc[:, 2].notna()]
sheet.iloc[:, 2] = sheet.iloc[:, 2].astype(str)
sheet = sheet[sheet.iloc[:, 2].apply(str.isnumeric)]

sheet = sheet[sheet.iloc[:, 6].apply(lambda x: (type(x) != datetime.datetime) and (type(x) != str))]
sheet = sheet[sheet.iloc[:, 3].notna()]

sheet.drop('index', axis=1, inplace=True)
sheet.reset_index(drop=True, inplace=True)

initial_indices = ["Hat", "Cihaz TTNr", "Cihaz Aile", "Toplam Adet"]
shifts = [1, 2, 3]
final_indices = [" - ".join([i, str(j)]) for i in shift_dates for j in shifts]
initial_indices.extend(final_indices)

sheet = sheet.set_axis(initial_indices, axis=1)

sheet["Cihaz TTNr"] = sheet["Cihaz TTNr"].astype(str)
pipes["Cihaz"] = pipes["Cihaz"].astype(str)

sheet = sheet.merge(pipes, left_on="Cihaz TTNr", right_on="Cihaz", how="inner")
sheet = sheet.merge(types, on="Boru", how="left")
sheet.drop("Cihaz", axis=1, inplace=True)
sheet.insert(3, 'Boru', sheet.pop('Boru'))
sheet["Cihaz TTNr"] = sheet["Cihaz TTNr"].astype("int64")

indices = ["Hat", "Cihaz TTNr", "Cihaz Aile", "Boru TTNr"]

df = pd.DataFrame(columns=pd.MultiIndex.from_product([shift_dates, shifts]),
                  index=range(sheet.shape[0]))
df = pd.concat([pd.DataFrame(columns=pd.MultiIndex.from_product([indices,
                                                                 ["" for _ in range(len(indices))]])), df], axis=1)

dates_df = df.iloc[:, 16:]
initial_df = df.swaplevel(axis=1, i=0, j=1).iloc[:, :16]
initial_df = initial_df.loc[:, ~initial_df.columns.duplicated()]
df = pd.concat([initial_df, dates_df], axis=1)

sheet.iloc[:, 4:-1] = sheet.iloc[:, 4:-1].apply(lambda x: x * sheet.iloc[:, -2], axis=0)

sheet.drop("Miktar", axis=1, inplace=True)
sheet.drop("Toplam Adet", axis=1, inplace=True)

df.iloc[:, :] = sheet.iloc[:, :]
df["Tip"] = sheet["Tip"]

type_df = df.swaplevel(axis=1, i=0, j=1).iloc[:, -1]
df = pd.concat([df.iloc[:, :-1], type_df], axis=1)
df.insert(4, ('', 'Tip'), df.pop(('', 'Tip')))

df = df.set_index(("", "Hat")).rename_axis(None, axis=0)

# Colors for excel formatting
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

df.iloc[:, 4:] = df.iloc[:, 4:].apply(pd.to_numeric, errors='coerce')

df_pivot = df.sort_index(key=lambda x: (x.to_series().str[6:].astype("int64")))
df_pivot = df_pivot.drop(columns=[('', 'Cihaz TTNr'), ('', 'Cihaz Aile'), ('', 'Tip')])

df_pivot.index = df_pivot.index.str.split(' ').str[1]
df_pivot = df_pivot.groupby([df_pivot.index, ("", "Boru TTNr")]).sum().sort_index(ascending=False)
df_pivot = df_pivot.reset_index(level=1, drop=False)
df_pivot.index = df_pivot.index.map(lambda x: f"Yalın {x}")
df_pivot["Tip"] = df_pivot.loc[:, ("", "Boru TTNr")].map(types.set_index("Boru")["Tip"])

swapped_types = df_pivot.iloc[:, -1].to_frame().swaplevel(axis=1, i=0, j=1).iloc[:, 0].to_frame()
df_pivot.insert(1, ('', 'Tip'), swapped_types)
df_pivot.drop(("Tip", ""), axis=1, inplace=True)

df_pivot.iloc[:, 2:] = df_pivot.iloc[:, 2:].applymap(lambda x: np.nan if x == 0 else x)


# format the Excel column dimensions
def general_excel_formatter(file_path: str):
    wb = openpyxl.load_workbook(file_path)

    ws1 = wb["Sheet1"]
    ws1.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws1)

    for col in range(ws1.min_column, ws1.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws1, min=col, max=col, width=12)

    #  change the height of all rows
    for row in range(ws1.min_row, ws1.max_row + 1):
        ws1.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws1, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws1, min=3, max=3, width=18)
    dim_holder['D'] = ColumnDimension(ws1, min=4, max=4, width=18)
    dim_holder['E'] = ColumnDimension(ws1, min=5, max=5, width=18)

    # add filter
    ws1.auto_filter.ref = "A2:E2"

    # highlight the version and date cells
    ws1['A1'].fill = redFill
    ws1['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1['B1'].fill = redFill
    ws1['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws1.column_dimensions = dim_holder
    wb.save(file_path)


version_value = version.iloc[0] + " - " + version.iloc[1]
update_date_value = update_date.iloc[0] + ":  " + update_date.iloc[1]


def pivot_excel_formatter(file_path: str):
    wb = openpyxl.load_workbook(file_path)

    ws2 = wb["Sheet2"]
    ws2.delete_rows(3)

    dim_holder = DimensionHolder(worksheet=ws2)

    for col in range(ws2.min_column, ws2.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws2, min=col, max=col, width=12)

    #  change the height of all rows
    for row in range(ws2.min_row, ws2.max_row + 1):
        ws2.row_dimensions[row].height = 20

    # change the size of B, C, and D columns
    dim_holder['B'] = ColumnDimension(ws2, min=2, max=2, width=18)
    dim_holder['C'] = ColumnDimension(ws2, min=3, max=3, width=18)

    # add filter
    ws2.auto_filter.ref = "A2:C2"

    # highlight the version and date cells
    ws2['A1'].fill = redFill
    ws2['A1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2['B1'].fill = redFill
    ws2['B1'].font = Font(color="FFFFFF", bold=True, size=11)

    ws2.column_dimensions = dim_holder
    wb.save(file_path)


# add the Excel version to the file
def excel_version(file_path: str):
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.cell(row=1, column=1).value = version_value
        ws.cell(row=1, column=2).value = str(update_date_value)
        ws.cell(row=2, column=1).value = "Hat"

        # center all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(file_path)


# write the dataframe to an Excel file
try:
    print(">>>\n>>> Conversion started...")

    if os.path.exists(f"../Data/Output/{excel_file}_output.xlsx"):
        wb = openpyxl.load_workbook(f"../Data/Output/{excel_file}_output.xlsx")

        if "Sheet1" not in wb.sheetnames:
            wb.create_sheet("Sheet1")
        if "Sheet2" not in wb.sheetnames:
            wb.create_sheet("Sheet2")

        wb.save(f"../Data/Output/{excel_file}_output.xlsx")

    with pd.ExcelWriter(f"../Data/Output/{excel_file}_output.xlsx", mode="w") as writer:
        df.to_excel(writer, sheet_name="Sheet1")
        df_pivot.to_excel(writer, sheet_name="Sheet2")

    print(">>> Conversion completed successfully!")
    print(">>> Excel Formatting started...")
    pivot_excel_formatter(file_path=f"../Data/Output/{excel_file}_output.xlsx")
    general_excel_formatter(file_path=f"../Data/Output/{excel_file}_output.xlsx")
    excel_version(file_path=f"../Data/Output/{excel_file}_output.xlsx")
    print(">>> Excel Formatting completed successfully!")
except Exception as e:
    print(e)
    print(">>> Conversion failed!")
finally:
    print(">>> Terminating...")
    exit(0)
